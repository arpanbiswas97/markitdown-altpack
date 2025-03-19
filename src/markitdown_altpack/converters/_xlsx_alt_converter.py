# markitdown-altpack
# Copyright (c) 2025 Arpan Biswas
# Licensed under the MIT License – see the LICENSE file in the root of the repository.

import datetime
import re
import typing as T

import cv2
import numpy as np
import pandas as pd
import xlrd
from markitdown import DocumentConverter, DocumentConverterResult, StreamInfo
from openpyxl import Workbook, load_workbook

XLSX_MIME_TYPE_PREFIXES = [
    "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
]
XLSX_FILE_EXTENSIONS = [".xlsx"]

XLS_MIME_TYPE_PREFIXES = [
    "application/vnd.ms-excel",
    "application/excel",
]
XLS_FILE_EXTENSIONS = [".xls"]

# Table Detection Heuristics Parameters
COMMON_VALUES_THRESHOLD: int = 6
MIN_VALUE_EXTRAPOLATION_THRESHOLD: int = 3
MAX_EXTRAPOLATION_FILL_THRESHOLD: int = 11


class ValueType:
    # enum for value type
    CATEGORICAL = 1
    NUMERIC = 2
    DATETIME = 3
    BOOLEAN = 4


# Heuristic to detect value cells
def find_value_cells(cells: np.ndarray) -> np.ndarray[ValueType]:
    # Find list of common values
    value_counts = {}
    for cell in cells.flatten():
        if cell is not None and isinstance(cell.value, str):
            value_counts[cell.value.strip()] = (
                value_counts.get(cell.value.strip(), 0) + 1
            )
    common_values = [v for v, c in value_counts.items() if c >= COMMON_VALUES_THRESHOLD]

    def map_value_type(cell) -> ValueType:
        if cell is None:
            return 0
        elif isinstance(cell.value, bool):
            return ValueType.BOOLEAN
        elif isinstance(cell.value, (int, float, complex)):
            return ValueType.NUMERIC
        elif isinstance(cell.value, datetime.datetime):
            return ValueType.DATETIME
        elif isinstance(cell.value, str):
            if cell.value.strip() in common_values:
                return ValueType.CATEGORICAL
            elif re.fullmatch(r"[\d\W]+", cell.value.strip()):
                return ValueType.NUMERIC
            else:
                return 0
        else:
            return 0

    # Find value cells
    return np.vectorize(map_value_type)(cells)


# Heuristic to break sheet into tables
def detect_bounding_boxes(
    pixel_matrix: np.ndarray,
) -> T.List[T.Tuple[int, int, int, int]]:
    """Detect bounding boxes"""

    # Convert to 0-255
    pixel_matrix = pixel_matrix.astype(np.uint8) * 255
    # Find contours
    contours, hierarchy = cv2.findContours(
        pixel_matrix, cv2.RETR_TREE, cv2.CHAIN_APPROX_SIMPLE
    )
    # Find bounding boxes
    bounding_boxes = [cv2.boundingRect(contour) for contour in contours]

    # Remove bounding boxes contained in other bounding boxes
    sorted_bboxes = sorted(bounding_boxes, key=lambda x: x[2] * x[3], reverse=True)
    filtered_bboxes = []
    for bbox in sorted_bboxes:
        if len(filtered_bboxes) == 0:
            filtered_bboxes.append(bbox)
        else:
            overlap = False
            for filtered_bbox in filtered_bboxes:
                if (
                    bbox[0] >= filtered_bbox[0]
                    and bbox[1] >= filtered_bbox[1]
                    and bbox[0] + bbox[2] <= filtered_bbox[0] + filtered_bbox[2]
                    and bbox[1] + bbox[3] <= filtered_bbox[1] + filtered_bbox[3]
                ):
                    overlap = True
                    break
            if not overlap:
                filtered_bboxes.append(bbox)

    return filtered_bboxes


def extract_tables_from_sheet(wb: Workbook, sheet_name: str) -> T.List[pd.DataFrame]:
    """Extract tables from a sheet."""

    ws = wb[sheet_name]
    if not ws.max_column or not ws.max_row:
        # Reset dimensions
        ws.calculate_dimension(force=True)
    num_rows = ws.max_row
    num_cols = ws.max_column
    pixel_matrix = np.full((num_rows, num_cols), 0, dtype=np.uint8)
    cells = np.full((num_rows, num_cols), None, dtype=object)

    # Iterate over all cells
    for r, row in enumerate(ws.iter_rows(max_row=num_rows, max_col=num_cols)):
        for c, cell in enumerate(row):
            if cell.value is not None and str(cell.value).strip() != "":
                pixel_matrix[r, c] = 1
                cells[r, c] = cell

    # Fill blanks in pixel matrix by extrapolating value cells
    value_cells = find_value_cells(cells)
    for c in range(value_cells.shape[1]):
        last_value_cell_type = None
        last_value_cell_index = None
        same_value_cell_count = 0
        for r in range(value_cells.shape[0] - 1, -1, -1):
            # Fill empty cells
            if cells[r, c] is not None:
                if (
                    last_value_cell_type is not None
                    and last_value_cell_index is not None
                    and last_value_cell_index - r > 1
                    and same_value_cell_count > MIN_VALUE_EXTRAPOLATION_THRESHOLD
                    and last_value_cell_index - r <= MAX_EXTRAPOLATION_FILL_THRESHOLD
                ):
                    pixel_matrix[r + 1 : last_value_cell_index, c] = 1
            # Update state
            if value_cells[r, c]:
                last_value_cell_index = r
                if last_value_cell_type == value_cells[r, c]:
                    same_value_cell_count += 1
                else:
                    last_value_cell_type = value_cells[r, c]
                    same_value_cell_count = 1
            elif cells[r, c] is not None:
                last_value_cell_type = None
                last_value_cell_index = None
                same_value_cell_count = 0

    # Detect bounding boxes
    bboxes = detect_bounding_boxes(pixel_matrix)

    # Sort bounding boxes by row index
    bboxes = sorted(bboxes, key=lambda x: x[1])

    # Create tables
    tables = []
    for bbox in bboxes:
        x, y, w, h = bbox
        # Define the range of cells to read
        skiprows = y
        usecols = list(range(x, x + w))
        nrows = h
        # Read the specific range of cells into a DataFrame
        pandas_wb = pd.ExcelFile(wb, engine="openpyxl")
        table = pandas_wb.parse(
            sheet_name=sheet_name,
            skiprows=skiprows,
            usecols=usecols,
            nrows=nrows,
        )
        tables.append(table)

    return tables


def convert_workbook_to_markdown(wb: Workbook) -> str:
    """Perform table detection and convert it to markdown"""
    sheets_md = []
    for sheet_name in wb.sheetnames:
        tables = extract_tables_from_sheet(wb, sheet_name)
        md = "# " + sheet_name + "\n\n"
        md += "\n\n<!-- TableBreak -->\n\n".join(
            [table.fillna("").to_markdown(index=False) for table in tables]
        )
        sheets_md.append(md)
    return "\n\n<!-- SheetBreak -->\n\n".join(sheets_md)


class XlsxAltConverter(DocumentConverter):
    """Alternate XLSX converter with table detection"""

    def accepts(
        self,
        file_stream: T.BinaryIO,
        stream_info: StreamInfo,
        **kwargs: T.Any,  # Options to pass to the converter
    ) -> bool:
        mimetype = (stream_info.mimetype or "").lower()
        extension = (stream_info.extension or "").lower()

        if extension in XLSX_FILE_EXTENSIONS:
            return True

        for prefix in XLSX_MIME_TYPE_PREFIXES:
            if mimetype.startswith(prefix):
                return True

    def convert(
        self,
        file_stream: T.BinaryIO,
        stream_info: StreamInfo,
        **kwargs: T.Any,  # Options to pass to the converter
    ) -> DocumentConverterResult:
        wb = load_workbook(
            file_stream, read_only=True, data_only=True, keep_links=False
        )

        markdown_text = convert_workbook_to_markdown(wb)

        return DocumentConverterResult(markdown=markdown_text)


class XlsAltConverter(DocumentConverter):
    """Alternate XLS converter with table detection"""

    def accepts(
        self,
        file_stream: T.BinaryIO,
        stream_info: StreamInfo,
        **kwargs: T.Any,  # Options to pass to the converter
    ) -> bool:
        mimetype = (stream_info.mimetype or "").lower()
        extension = (stream_info.extension or "").lower()

        if extension in XLS_FILE_EXTENSIONS:
            return True

        for prefix in XLS_MIME_TYPE_PREFIXES:
            if mimetype.startswith(prefix):
                return True

    def convert(
        self,
        file_stream: T.BinaryIO,
        stream_info: StreamInfo,
        **kwargs: T.Any,  # Options to pass to the converter
    ) -> DocumentConverterResult:

        # Load Xls and convert to Xlsx
        book_xls = xlrd.open_workbook(file_contents=file_stream.read())
        wb = Workbook()
        sheet_names = book_xls.sheet_names()
        for sheet_index, sheet_name in enumerate(sheet_names):
            sheet_xls = book_xls.sheet_by_name(sheet_name)
            if sheet_index == 0:
                sheet_xlsx = wb.active
                sheet_xlsx.title = sheet_name
            else:
                sheet_xlsx = wb.create_sheet(title=sheet_name)

            for row in range(0, sheet_xls.nrows):
                for col in range(0, sheet_xls.ncols):
                    sheet_xlsx.cell(row=row + 1, column=col + 1).value = (
                        sheet_xls.cell_value(row, col)
                    )

        markdown_text = convert_workbook_to_markdown(wb)

        return DocumentConverterResult(markdown=markdown_text)
